import openpyxl
import os,json,sys
import datetime
from util.handle_log import Logger
base_path=os.path.dirname(os.path.dirname(__file__))
file_path = base_path+"/data/Test_case.xlsx"
now = datetime.datetime.now().strftime('%Y-%m-%d %H_%M_%S')
new_file_path=base_path+"\report\%s.xlsx" % now
sys.path.append(base_path)
logger=Logger(logger="HandleExcel").getlog()

class HandleExcel():
    """
    定义一个操作Excel数据表的类
    """
    def loadExcel(self):
        """
        打开Excel数据表
        """
        open_excel=openpyxl.load_workbook(file_path)
        return open_excel

    def get_sheet_names(self):
        """
        获取sheetname集合
        """
        sheet_names=self.loadExcel().sheetnames
        return sheet_names

    def getSheetData(self,index=None):
        """
        加载Excel表的sheet
        """
        sheet_names=self.loadExcel().sheetnames
        #print(sheet_names)
        if index==None:
            index=0
        data=self.loadExcel()[sheet_names[index]]
        return data

    def getCellValue(self,row,cols,index=None):
        """
        获取某一单元格的内容
        """
        data=self.getSheetData(index).cell(row=row,column=cols).value
        return data

    def getRows(self,index=None):
        """
        获取行数
        """
        rows=self.getSheetData(index).max_row
        return rows

    def getRowValue(self,row,index=None):
        """
        获取某一行的内容
        """
        row_list=[]
        for i in self.getSheetData(index)[row]:
            #print(i)  #<Cell 'Sheet1'.A2>为每个单元格的对象
            row_list.append(i.value)
        return row_list

    def getColValue(self,index=None,col=None):
        """
        获取某一列的内容
        """
        col_list=[]
        if col==None:
            col="A"
        for i in self.getSheetData(index)[col]:
            col_list.append(i.value)
        return col_list

    def getRowsNumber(self,data,index=None):
       """
       获取行号
       """
       num=1
       cols_data=self.getColValue(index) # 返回该列所有数据，已列表格式存在
       #print(cols_data)
       for col_data in cols_data: # 遍历返回的列数据
           if col_data==data: # 通过传入的单元格数据，进行列数据比对，相同则返回当前num，不相同则继续比对，同时num+1
            return num
           num=num+1
       return num

    def writeData(self,row,col,value,index=None):
        """
        写入数据
        """
        wb=self.loadExcel() # 打开excel表
        work_sheet=wb[self.get_sheet_names()[index]] # 定位需求操作的sheet
        work_sheet.cell(row=row,column=col,value=value) # 通过row,col，定位到单元格，并写入value
        wb.save(file_path) # 将写入的数据保存到file_path

    def getExcelData(self,index=None):
        """
        获取Excel的所有数据
        """
        data_list=[]
        for i in range(self.getRows(index)-1):
            data_list.append(self.getRowValue(i+2,index))
        return data_list

    def new_excel(self):
        """
        新建Excel表
        """
        open_excel =openpyxl.Workbook()
        open_excel.save(new_file_path)
        return "新建Excel成功，文件名路径：%s" % new_file_path

handle_excel=HandleExcel()

if __name__ == '__main__':
    handle_excel=HandleExcel()
    data=handle_excel.getCellValue(2,7)
    print(data)
    if data.find("####")!=-1:
        print(data.replace("####","daad"))





